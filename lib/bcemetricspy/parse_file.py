import os
import random
import openpyxl
import numpy as np

def matrix_bids(wb, sheet1, sheet2, nb_bidders, bid_threshold=float("inf")):
    index = 0
    id_to_index = {}
    # Find the number of tracts in the data
    nb_bidders_per_tract = []
    acre_per_tract = []
    for row in range(2, sheet1.max_row + 1):
        nb_bidders_per_tract.append(int(sheet1['K' + str(row)].value))
        acre_per_tract.append(int(sheet1['G' + str(row)].value))
        tract_id = int(sheet1['B' + str(row)].value)
        id_to_index[tract_id] = index
        index += 1

    # Compute two lists:
    #- list_id_bids is a list of all [tract number,bid] pairs present in the data
    #- nb_bidders_per_tract computes the number of bidders for each tract

    list_id_bids = []
    for row in range(2, sheet2.max_row + 1):
        bid = float(sheet2['F' + str(row)].value)
        tract_id = int(sheet2['B' + str(row)].value)
        index = id_to_index[tract_id]
        acre = acre_per_tract[index]
        bid_per_acre = bid / acre
        list_id_bids.append([tract_id, bid_per_acre])

    # Compute lists of all bids from 2-player auctions

    list_bids = []
    for index in range(nb_bidders):
        list_bids.append([])

    # Create two lists: list_bids_p1 contains the bids of player 1, list_bids_p2 contains the bids of player 2.
    # Bids are ordered in the same order as auctions. Only auctions with 2
    # players are considered
    index = 0
    random.seed(1526)
    while index <= len(list_id_bids) - 1:
        tract = id_to_index[list_id_bids[index][0]]
        if nb_bidders_per_tract[tract] == nb_bidders:
            # Randomize the bidders here
            X = list(range(nb_bidders))
            random.shuffle(X)
            for bidder in range(nb_bidders):
                list_bids[bidder].append(list_id_bids[index + X[bidder]][1])
            index += nb_bidders
        else:
            index += 1

    list_bids = np.asarray(list_bids)
    list_bids_thres = []
    for bidder in range(nb_bidders):
        list_bids_thres.append([])
    for auction in range(len(list_bids[0])):
        if max(list_bids[:, auction]) <= bid_threshold:
            for bidder in range(nb_bidders):
                list_bids_thres[bidder].append(list_bids[bidder, auction])

    return list_bids_thres

def bins_from_matrix(list_bids,max_bid):   
    nb_bids = len(list_bids[0])
    nb_bidders = len(list_bids)

    #Renormalize the bids to be between 0 and max_bid
    max_real_bid = max([max(line) for line in list_bids])
    size_bin = max_real_bid / max_bid
    for index in range(nb_bids):
        for bidder in range(nb_bidders):
            list_bids[bidder][index] = list_bids[bidder][index] / float(size_bin)

    # Compute the joint distribution of bids for all players
    # Round values to the nearest integer bin to decide on a bin

    bid_pdf = {}
    for index in range(nb_bids):
        tple = [int(round(list_bids[bidder][index])) for bidder in range(nb_bidders)]
        tple = tuple(tple)
        bid_pdf[tple] = 0

    for index in range(nb_bids):
        tple = [int(round(list_bids[bidder][index])) for bidder in range(nb_bidders)]
        tple = tuple(tple)
        bid_pdf[tple] += 1 / float(nb_bids)

    return bid_pdf

def bins_from_data(wb, sheet1, sheet2, nb_bidders, max_bid, bid_threshold = float("inf")):
    bid_tr = bid_threshold
    list_bids = matrix_bids(wb, sheet1, sheet2, nb_bidders, bid_threshold = bid_tr)
    bid_pdf = bins_from_matrix(list_bids, max_bid)

    return bid_pdf

if __name__ == '__main__':

    wb = openpyxl.load_workbook(os.path.abspath('Dataset.xlsx'))
    sheet1 = wb.get_sheet_by_name('Tract79')
    sheet2 = wb.get_sheet_by_name('Trbid79')

    # Testing we have the right number of auctions for 2 bidders
    nb_bidders = 2
    list_bids = matrix_bids(wb, sheet1, sheet2, nb_bidders)
    nb_bids = len(list_bids[0])

    if nb_bids == 584:
        nbbid2_test = 1
    else:
        nbbid2_test = 0

    if nbbid2_test == 1:
        print("Nb_auctions_2 test: passed")
    else:
        print("Nb_auctions_2 test: failed")
        
    # Testing we have the right number of auctions for 3 bidders
    nb_bidders = 3
    list_bids = matrix_bids(wb, sheet1, sheet2, nb_bidders)
    nb_bids = len(list_bids[0])

    if nb_bids == 330:
        nbbid3_test = 1
    else:
        nbbid3_test = 0

    if nbbid3_test == 1:
        print("Nb_auctions_3 test: passed")
    else:
        print("Nb_auctions_3 test: failed")

    # Testing the seeds work
    nb_bidders=2
    list_bids1 = matrix_bids(wb, sheet1, sheet2, nb_bidders)
    list_bids2 = matrix_bids(wb, sheet1, sheet2, nb_bidders)
    if list_bids1 == list_bids2:
        seed_test = 1
    else:
        seed_test = 0

    if seed_test == 1:
        print("Seed test: passed")
    else:
        print("Seed test: failed")

    # Testing we get the same bid/acre as in the data for the first 3 and last
    # 3 auctions of size 2
    nb_bidders = 2
    list_bids = matrix_bids(wb, sheet1, sheet2, nb_bidders)
    real_bids = [[5559969.0 / 2500.0, 1754865.0 / 2500.0],
                 [1219718.0 / 5000.0, 579106.0 / 5000.0],
                 [1316149.0 / 2500.0, 142474.0 / 2500.0],
                 [35772095.0 / 1568.0, 20705206.0 / 1568.0],
                 [2712131.0 / 5760.0, 1916748.0 / 5760.0],
                 [1542811.0 / 634.0, 581511.0 / 634.0]]

    fetched_bids = [[list_bids[0][0], list_bids[1][0]],
                    [list_bids[0][1], list_bids[1][1]],
                    [list_bids[0][2], list_bids[1][2]],
                    [list_bids[0][-3], list_bids[1][-3]],
                    [list_bids[0][-2], list_bids[1][-2]],
                    [list_bids[0][-1], list_bids[1][-1]]]

    bid2_test = 1
    for index in range(len(real_bids)):
        real_bids[index].sort()
        fetched_bids[index].sort()
        if real_bids[index] != fetched_bids[index]:
            bid2_test = 0
            break

    if bid2_test == 1:
        print("Bid2 test: passed")
    else:
        print("Bid2 test: failed")

    # Testing we get the same bid/acre as in the data for the first and last
    # auctions of size 3
    nb_bidders = 3
    list_bids = matrix_bids(wb, sheet1, sheet2, nb_bidders)

    real_bids = [[3822479.0/5000.0,1742703.0/5000.0,701946.0/5000.0],
                 [4616985.0/4994.0,960282.0/4994.0,190848.0/4994.0]]

    fetched_bids = [[list_bids[0][0], list_bids[1][0],list_bids[2][0]],
                    [list_bids[0][-1], list_bids[1][-1],list_bids[2][-1]]]

    bid3_test = 1
    for index in range(len(real_bids)):
        real_bids[index].sort()
        fetched_bids[index].sort()
        if real_bids[index] != fetched_bids[index]:
            bid3_test = 0
            break

    if bid3_test == 1:
        print("Bid3 test: passed")
    else:
        print("Bid3 test: failed")

    # Testing bins_from_list outputs a probability distribution
    max_bid=50
    bid_pdf=bins_from_matrix(list_bids,max_bid)
    proba_vect=[bid_pdf[key] for key in bid_pdf.keys()]
    s=sum(proba_vect)
    min_prob=min(proba_vect)

    if abs(s-1) <= 10**(-6) and min_prob >= 0.0:
        proba_test=1
    else:
        proba_test=0

    if proba_test == 1:
        print("Proba test: passed")
    else:
        print("Proba test: failed")


    sum_tests = nbbid2_test + nbbid3_test + seed_test + bid2_test + bid3_test + proba_test
    if sum_tests == 6:
        print("==> All tests passed!")
    else:
        print("==> At least one test failed")
